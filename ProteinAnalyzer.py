"""
    Created by Vincent Wu on 1/18/17:
    This program analyzes the hydrophobic and hydrophillic blocks of a protein amino acid sequence
"""
#All neccesary imports
import matplotlib
import random
import math
import json
import os.path
import numpy as np
matplotlib.use('TkAgg')
from numpy import arange, sin, pi
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
from matplotlib import style
# implement the default mpl key bindings
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure
import sys
if sys.version_info[0] < 3:
    import Tkinter as Tk
    from Tkinter import ttk
    from Tkinter import Image
    import Tkinter.filedialog
    import Tkinter.messagebox
    print(sys.version_info[0], "hi")
else:
    import tkinter as Tk
    from tkinter import ttk
    #from tkinter import Image
    from tkinter.colorchooser import *
    import tkinter.filedialog
    import tkinter.messagebox
    print(sys.version_info[0])
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw
from openpyxl.drawing.image import Image as xlImage
#global variables
VERSION = "v1.2"
FILE = "2zc1.fasta"
PH = 7
HYDROLIMIT = 0
NUMCUT = 1
GRAPH1 = 1
GRAPH2 = 2
ALT = True
STYLE = "bmh"
COLOR1 = '#4D4D4D'
COLOR2 = '#5DA5DA'
COLOR3 = '#F15854'
COLOR4 = '#DECF3F'
COLOR5 = '#60BD68'
COLOR6 = '#F17CB0'
COLOR7 = '#B276B2'
COLOR8 = '#FAA43A'
DCOLOR1 = '#353535'
DCOLOR2 = '#4a84ae'
DCOLOR3 = '#c04643'
DCOLOR4 = '#b1a532'
DCOLOR5 = '#4c9753'
DCOLOR6 = '#c0638c'
DCOLOR7 = '#8e5e8e'
DCOLOR8 = '#c8832e'
LCOLOR1 = '#828282'
LCOLOR2 = '#8dc0e5'
LCOLOR3 = '#f58a87'
LCOLOR4 = '#e7dd78'
LCOLOR5 = '#8fd095'
LCOLOR6 = '#f5a3c7'
LCOLOR7 = '#c99fc9'
LCOLOR8 = '#fbbf75'
COLORARRAY = [COLOR1, COLOR2, COLOR3, COLOR4, COLOR5, COLOR6, COLOR7, COLOR8]
DCOLORARRAY = [DCOLOR1, DCOLOR2, DCOLOR3, DCOLOR4, DCOLOR5, DCOLOR6, DCOLOR7, DCOLOR8]
LCOLORARRAY = [LCOLOR1, LCOLOR2, LCOLOR3, LCOLOR4, LCOLOR5, LCOLOR6, LCOLOR7, LCOLOR8]
STYLE = "seaborn-muted"

#amino acid class
class aminoAcid:
	def __init__(self, name, hphob2, hphob7, alt_hphob):
		self.name =  name
		self.hphob2 = hphob2
		self.hphob7= hphob7
		self.alt_hphob = alt_hphob
#instantiating all amino acid objects
ALA = aminoAcid("A", 47, 41, 1.8)
ARG = aminoAcid("R", -26, -14, -4.5)
ASN = aminoAcid("N", -41, -28, -3.5)
ASP = aminoAcid("D", -18, -55, -3.5)
CYS = aminoAcid("C",  52, 49, 2.5)
GLU = aminoAcid("E", 8, -31, -3.5)
GLN = aminoAcid("Q", -18, -10, -3.5)
GLY = aminoAcid("G", 0, 0, -0.4)
HIS = aminoAcid("H", -42, 8, -3.2)
HYP = aminoAcid("O", None, None, None)
ILE = aminoAcid("I", 100, 99, 4.5)
LEU = aminoAcid("L", 100, 97, 3.8)
LYS = aminoAcid("K", -37, -23, -3.9)
MET = aminoAcid("M", 74, 74, 1.9)
PHE = aminoAcid("F", 92, 100, 2.8)
PRO = aminoAcid("P", -46, -46, -1.6)
GLP = aminoAcid("U", None, None, None)
SER = aminoAcid("S", -7, -5, -0.8)
THR = aminoAcid("T", 13, 13, -0.7)
TRP = aminoAcid("W", 84, 97, -0.9)
TYR = aminoAcid("Y", 49, 63, -1.3)
VAL = aminoAcid("V", 79, 76, 4.2)
aminoAcidList = [ALA, ARG, ASN, ASP, CYS, GLU, GLN, GLY, HIS, HYP, ILE, LEU, LYS, 
MET, PHE, PRO, GLP, SER, THR, TRP, TYR, VAL]
#given the letter name, returns the correct amino acid object
def getAminoAcid(name):
	for acid in aminoAcidList:
		if acid.name == name:
			return acid
	assert False, "Amino Acid %s not found!" %name
def parseFile(filename):
	try:
		file = open(filename, 'r')
	except:
		errorMessage("Filename not found!", 180)
		return
	sequenceString = file.read().replace(" ","").replace("\n", "")
	file.close()
	sequenceList = list(sequenceString)
	remove = False
	removeRest1 = False
	removeRest2 = False
	codecount = 0
	for char in sequenceList[:]:
		print(char)
		if removeRest2:
			sequenceList.remove(char)
		elif char == '>':
			if removeRest1:
				removeRest2 = True
			removeRest1 = True
			sequenceList.remove(char)
			remove = True
		elif removeRest1 == False:
			break
		elif remove:
			if codecount == 0:
				if char == 'N':
					codecount = 1
			elif codecount == 1:
				if char == 'C':
					codecount = 2
			elif codecount == 2:
				if char == 'E':
					remove = False
					print("false!")
			else:
				codecount = 0
			sequenceList.remove(char)
		#print("remove: ", remove)
		#print("removeRest1: ", removeRest1)
		#print("removeRest2: ", removeRest2)
		#print("codecount: ", codecount)

	file = open(filename, 'w')
	file.write(''.join(sequenceList))
	file.close()
	#print(sequenceList)
	return sequenceList
class Application(ttk.Frame):
	def __init__(self, master = None):
		ttk.Frame.__init__(self, master)
		self.pack()
		self.initialize()
	def initialize(self):
		self.graphExists = False
		self.altGraphExists = False
		self.genFrame = Tk.Frame(master = root)
		self.genFrame.pack(side = Tk.LEFT)
		self.inputsFrame = ttk.Frame(master = self.genFrame)
		self.inputsFrame.pack(side = Tk.TOP, padx= 0, pady = 6)
		self.fileFrame = Tk.Frame(master = self.inputsFrame, bg = "#bbc3cc")
		self.fileFrame.pack(side = Tk.TOP, pady = 0, padx = 0)
		self.fileLabel = Tk.Label(master = self.fileFrame, text = "File Name:              ", bg = "#bbc3cc")
		self.fileLabel.pack(side = Tk.LEFT, padx = 3, pady = 3)
		self.fileTkVar = Tk.StringVar()
		self.fileEntry = ttk.Entry(master = self.fileFrame, width = 10, textvariable = self.fileTkVar)
		self.fileEntry.pack(side = Tk.LEFT, padx = 4, pady = 5)
		self.fileTkVar.set(FILE)
		self.numCutFrame = Tk.Frame(master = self.inputsFrame, bg = "#9fbfdf")
		self.numCutFrame.pack(side = Tk.TOP, pady = 0)
		self.numCutLabel = Tk.Label(master = self.numCutFrame, text = "Number of Cutoffs:        ", bg = "#9fbfdf")
		self.numCutLabel.pack(side = Tk.LEFT, padx = 4, pady = 3)
		self.numCutTkVar = Tk.IntVar()
		self.numCutCombobox = ttk.Combobox(master = self.numCutFrame, values = [1,2,3,4], width = 2, textvariable = self.numCutTkVar,
			 state = "readonly")
		self.numCutCombobox.bind("<<ComboboxSelected>>", lambda e: self.updateCutoff(self))
		self.numCutTkVar.set(NUMCUT)
		self.currNumCutoffs = 1
		self.numCutCombobox.pack(side = Tk.LEFT, padx = 4, pady = 3)
		self.pHFrame = Tk.Frame(master = self.inputsFrame, bg = "#bbc3cc")
		self.pHFrame.pack(side = Tk.TOP, pady = 0) 
		self.pHLabel = Tk.Label(master = self.pHFrame, text = "pH:                                    ", bg = "#bbc3cc")
		self.pHLabel.pack(side = Tk.LEFT, padx = 3, pady = 3)
		self.pHTkVar = Tk.IntVar()
		self.pHCombobox = ttk.Combobox(master = self.pHFrame, values = [2,7], textvariable = self.pHTkVar, state = "readonly", width = 2)
		self.pHFrame = Tk.Frame(master = self.inputsFrame, bg = "#bbc3cc")
		self.pHTkVar.set(PH)
		self.pHCombobox.pack(side = Tk.LEFT, padx = 5, pady = 3)
		self.graph1Frame = Tk.Frame(master = self.inputsFrame, bg = "#9fbfdf")
		self.graph1Frame.pack(side = Tk.TOP, pady = 0) 
		self.graph1Label = Tk.Label(master = self.graph1Frame, text = "Graph 1:                           ", bg = "#9fbfdf#")
		self.graph1Label.pack(side = Tk.LEFT, padx = 4, pady = 3)
		self.graph1TkVar = Tk.IntVar()
		self.graph1Combobox = ttk.Combobox(master = self.graph1Frame, values = [1,2,3,4,5], textvariable = self.graph1TkVar, state = "readonly", width = 2)
		self.graph1TkVar.set(GRAPH1)
		self.graph1Combobox.pack(side = Tk.LEFT, padx = 5, pady = 3)
		self.graph2Frame = Tk.Frame(master = self.inputsFrame, bg = "#bbc3cc")
		self.graph2Frame.pack(side = Tk.TOP, pady = 0) 
		self.graph2Label = Tk.Label(master = self.graph2Frame, text = "Graph 2:                           ", bg = "#bbc3cc")
		self.graph2Label.pack(side = Tk.LEFT, padx = 4, pady = 3)
		self.graph2TkVar = Tk.IntVar()
		self.graph2Combobox = ttk.Combobox(master = self.graph2Frame, values = [1,2,3,4,5], textvariable = self.graph2TkVar, state = "readonly", width = 2)
		self.graph2TkVar.set(GRAPH2)
		self.graph2Combobox.pack(side = Tk.LEFT, padx = 5, pady = 3)
		self.cutFrameList = []
		self.cutFrameList.append(self.numCutFrame)
		self.hydroLimitFrame = Tk.Frame(master = self.inputsFrame, bg = "#9fbfdf")
		self.hydroLimitFrame.pack(side = Tk.TOP, pady = 0)
		self.hydroLimitLabel = Tk.Label(master = self.hydroLimitFrame, text = "Hydrophobicity Cutoff 1:", bg = "#9fbfdf")
		self.hydroLimitLabel.pack(side = Tk.LEFT, padx = 0, pady = 3)
		self.hydroLimitTkVarList = []
		hydroLimitTkVar = Tk.IntVar()
		self.hydroLimitTkVarList.append(hydroLimitTkVar)
		self.hydroLimitEntry = ttk.Entry(master = self.hydroLimitFrame, width = 5, textvariable = hydroLimitTkVar)
		hydroLimitTkVar.set(HYDROLIMIT)
		self.hydroLimitEntry.pack(side = Tk.LEFT, padx = 5, pady = 5)
		self.currColor = "blue"
		self.buttonFrame = Tk.Frame(master = self.genFrame)
		self.buttonFrame.pack(side = Tk.TOP)
		self.bRowFrame1 = Tk.Frame(master = self.buttonFrame)
		self.bRowFrame1.pack(side = Tk.TOP)
		self.bRowFrame2 = Tk.Frame(master = self.buttonFrame)
		self.bRowFrame2.pack(side = Tk.TOP)
		self.analyzeButton = ttk.Button(master = self.bRowFrame1, text = "Analyze", width = 8, command = lambda: self.analyze())
		self.analyzeButton.pack(side = Tk.LEFT, pady = 6)
		self.visualizeButton = ttk.Button(master = self.bRowFrame1, text = "Visualize", width = 8, command = lambda: self.visualize())
		self.visualizeButton.pack(side = Tk.LEFT, padx = 5)
		self.exportButton = ttk.Button(master = self.bRowFrame2, text = "Export", width = 8, command = lambda: self.exportData())
		self.exportButton.pack(side = Tk.LEFT, padx = 0)
		self.graphButton = ttk.Button(master = self.bRowFrame2, text = "Graph", width = 8, command = lambda:self.graph())
		self.graphButton.pack(side = Tk.LEFT, padx = 5)
		self.canvasFrame = Tk.Frame(master = root)
		self.canvasFrame.pack(side = Tk.LEFT, fill = Tk.BOTH, expand = 1)
		self.startCanvas = Tk.Canvas(master = self.canvasFrame, width = 550, height = 330, bd = 0, relief = "ridge", highlightthickness = 0)
		self.startCanvas.pack(side = Tk.TOP, fill = Tk.BOTH, expand = 1)
		self.startCanvas.configure(bg = "#dde8f1")
		self.startCanvas.create_text(280, 330 / 3 + 30, text = "Welcome to ProteinAnalyzer! You must have a .txt file with your desired sequence\n " + 
			"in the same directory as the executable. To analyze, type in your .txt filename, \n" + " adjust the inputs, and press Analyze.")
		#self.photo = Tk.PhotoImage(file = "gfp.png")
		#self.startCanvas.create_image(546 / 2, 326 / 2, image = self.photo)
		root.bind("<Return>", lambda e: self.key())
		adjust(root, 0.4)
	def key(self):
		self.exportData()

	def updateCutoff(event, self):
		newNumCuts = int(self.numCutTkVar.get())
		print("currNumCuts:", self.currNumCutoffs)
		print("newNumCuts: ", newNumCuts)
		
		if newNumCuts < self.currNumCutoffs:
			frameID = self.currNumCutoffs
			while frameID != newNumCuts:
				self.cutFrameList[frameID - 1].destroy()
				del self.cutFrameList[-1]
				if self.currColor == "blue":
					self.currColor = "gray"
				else:
					self.currColor = "blue"
					print("Removed ", frameID)
					print("\n")
				frameID -= 1
		elif newNumCuts > self.currNumCutoffs:
			for frameID in range(self.currNumCutoffs + 1, newNumCuts + 1):
				if self.currColor == "blue":
					hydroLimitFrame = Tk.Frame(master = self.inputsFrame, bg = "#bbc3cc")
					hydroLimitFrame.pack(side = Tk.TOP, pady = 0)
					hydroLimitLabel = Tk.Label(master = hydroLimitFrame, text = "Hydrophobicity Cutoff %i:" %frameID, bg = "#bbc3cc")
					hydroLimitLabel.pack(side = Tk.LEFT, padx = 0, pady = 3)
					hydroLimitTkVar = Tk.IntVar()
					self.hydroLimitTkVarList.append(hydroLimitTkVar)
					hydroLimitEntry = ttk.Entry(master = hydroLimitFrame, width = 5, textvariable = hydroLimitTkVar)
					hydroLimitTkVar.set(HYDROLIMIT)
					hydroLimitEntry.pack(side = Tk.LEFT, padx = 5, pady = 5)
					self.cutFrameList.append(hydroLimitFrame)
					self.currColor = "gray"
				else:
					hydroLimitFrame = Tk.Frame(master = self.inputsFrame, bg = "#9fbfdf")
					hydroLimitFrame.pack(side = Tk.TOP, pady = 0)
					hydroLimitLabel = Tk.Label(master = hydroLimitFrame, text = "Hydrophobicity Cutoff %i:" %frameID, bg = "#9fbfdf")
					hydroLimitLabel.pack(side = Tk.LEFT, padx = 0, pady = 3)
					hydroLimitTkVar = Tk.IntVar()
					self.hydroLimitTkVarList.append(hydroLimitTkVar)
					hydroLimitEntry = ttk.Entry(master = hydroLimitFrame, width = 5, textvariable = hydroLimitTkVar)
					hydroLimitTkVar.set(HYDROLIMIT)
					hydroLimitEntry.pack(side = Tk.LEFT, padx = 5, pady = 5)
					self.cutFrameList.append(hydroLimitFrame)
					self.currColor = "blue"
				print("\n")
		else:
			return
			print("\n")
			#assert False, "newNumcuts equals currNumCuts, should not happen!"
		self.currNumCutoffs = newNumCuts

	def analyze(self):
		filename = self.fileTkVar.get() + ".txt"
		#print(filename)
		self.sequenceList = parseFile(filename)
		#print(sequenceList)
		self.sortedAminoList = self.sortAminoAcids()
		self.plotDistributions()
	def visualize(self):
		filename = self.fileTkVar.get() + ".txt"
		#print(filename)
		self.sequenceList = parseFile(filename)
		self.sortedAminoList = self.sortAminoAcids()
		#Toplevel parameters
		top = Tk.Toplevel()
		width = 1100
		height = 60
		top.focus_force()
		#top.grab_set()
		top.wm_title("Protein Visualization")
		top.geometry("%dx%d%+d%+d" % (width, height, 250, 125))
		adjust(top, 0.72)
		self.proteinCanvas = Tk.Canvas(master = top, width = width, height = height)
		self.proteinCanvas.pack(fill = Tk.BOTH, expand = 1)
		#self.proteinImage.paste(trans, (0, 0))
		xlength = int(width / len(self.sequenceList))
		print("side: ", xlength)
		ulx = (width - len(self.sequenceList)*xlength)/2
		ulx2 = 0
		ylength = 40
		uly = height/2 - ylength/2
		uly2 = 0
		prevAmino = -1
		consecutive = 0
		self.proteinImage = Image.new("RGB", (xlength*self.proteinLength, ylength), "#FFFFFF")
		draw = ImageDraw.Draw(self.proteinImage)
		for amino in self.sortedAminoList:
			color = COLORARRAY[amino]
			dcolor = DCOLORARRAY[amino]
			lcolor = LCOLORARRAY[amino]
			if prevAmino == amino:
				consecutive += 1
				self.proteinCanvas.create_rectangle(ulx - consecutive*xlength, uly, ulx + xlength, uly + ylength, fill = color, width = 0, activefill = lcolor)
				#draw.rectangle(((ulx2 - consecutive*xlength, uly2),(ulx2 + xlength, uly2 + ylength)), fill = color)
			else:
				consecutive = 0
				self.proteinCanvas.create_rectangle(ulx, uly, ulx + xlength, uly + ylength, fill = color, width = 0, activefill = lcolor)
				#draw.rectangle(((ulx2,uly2), (ulx2 + xlength, uly2 + ylength)), fill = color)
			draw.rectangle(((ulx2,uly2), (ulx2 + xlength, uly2 + ylength)), fill = color)
			prevAmino = amino
			ulx += xlength
			ulx2 += xlength
	#returns an array of the hydrophobicities of eahc amino acid in protein chain
	def getHydroArray(self):
		hydroList = []
		for aminoAcid in self.sequenceList:
			aminoAcidObj = getAminoAcid(aminoAcid)
			hydroList.append(aminoAcidObj.alt_hphob)
		#print(self.hydroList)
		return hydroList
	def graph(self):
		filename = self.fileTkVar.get() + ".txt"
		#print(filename)
		self.sequenceList = parseFile(filename)
		self.hydroArray = self.getHydroArray()
		index = []
		for i in range(1,len(self.hydroArray) + 1):
			index.append(i)
		style.use(STYLE)
		font = {'fontname':'Helvetica'}
		if not self.graphExists	:
			self.startCanvas.destroy()
			self.plotFigure = Figure(figsize=(5.5, 3.3), dpi=100, facecolor = "#dde8f1")
			self.subplot1 = self.plotFigure.add_subplot(111)
			self.subplot1.tick_params(labelsize = 7)
		elif not self.altGraphExists:
			self.plotFigure = Figure(figsize=(5.5, 3.3), dpi=100, facecolor = "#dde8f1")
			self.subplot1 = self.plotFigure.add_subplot(111)
			self.subplot1.tick_params(labelsize = 7)
		if self.altGraphExists:
			self.subplot1.clear()
		#self.subplot1.plot(index, self.hydroArray)
		npHydroArray = np.asarray(self.hydroArray)
		zeroArray = np.zeros(len(self.hydroArray))
		self.subplot1.fill_between(index, zeroArray, npHydroArray, where=npHydroArray<=zeroArray, linewidth = 0, facecolor = '#4D4D4D', interpolate = True)
		self.subplot1.fill_between(index, zeroArray, npHydroArray, where=npHydroArray>zeroArray, linewidth = 0,  facecolor = '#000080', interpolate = True)
		self.subplot1.set_xlabel("Protein Index", labelpad=5, fontsize = 8, **font)
		self.subplot1.set_ylabel("Hydropathic Index", labelpad=3, fontsize = 7, **font)
		if not self.graphExists:
			self.canvas = FigureCanvasTkAgg(self.plotFigure, master = self.canvasFrame)
		else:
			self.canvas.draw()
		self.canvas.show()
		if not self.graphExists:
			#self.toolbar = NavigationToolbar2TkAgg(self.canvas, root)
			#self.toolbar.update()
			self.canvas._tkcanvas.pack(side = Tk.BOTTOM, fill = Tk.BOTH, expand = 1)
			self.canvas.get_tk_widget().configure(bg = "black", bd = 0, selectbackground = "black")
			self.canvas.get_tk_widget().pack(side = Tk.BOTTOM, fill = Tk.BOTH, expand = 1)
			self.graphExists = True
		self.altGraphExists = True

	def exportData(self):
		self.analyze()
		self.histDataList = []
		class histData:
			def __init__(self, name, data):
				self.name =  name
				self.data = data
		for acidID in range(0,self.numCutTkVar.get()+1):
			binwidth = 1
			data, bins = np.histogram(self.getHistogramData(acidID),bins=range(1, max(self.getHistogramData(acidID)) + binwidth + 1, binwidth))
			print("data: ", data)
			print("bins: ", bins)
			data = data / self.proteinLength
			if self.numCuts == 1:
				if acidID == 0:
					name = "Hydrophilic Block Size"
				else: 
					name = "Hydrophobic Block Size"
			else:
				if acidID == self.numCuts:
					name = str(self.limitList[acidID - 1]) + " to " + str(100) + " Block Size"
				elif acidID == 0:
					name = str(-100) + " to " + str(self.limitList[acidID]) + " Block Size"
				else:
					name = str(self.limitList[acidID - 1]) + " to " + str(self.limitList[acidID]) + " Block Size"
			currhistData = histData(name, data)
			#print("histdata: ", data)
			self.histDataList.append(currhistData)
		print("histDataList: ", self.histDataList)
		wb = Workbook()
		ws = wb.active
		colCount = 1
		maxRow = 0
		for histData in self.histDataList:
			ws.cell(row = 1, column = colCount, value = histData.name)
			ws.column_dimensions[get_column_letter(colCount)].width = len(histData.name) - 1
			ws.cell(row = 1, column = colCount + 1, value = "Normalized Frequency")
			ws.column_dimensions[get_column_letter(colCount +1)].width = 20
			columnCount = 1
			maxRow = max(len(histData.data), maxRow)
			for column in ws.iter_cols(min_row = 2, max_row = len(histData.data) + 1, min_col = colCount, max_col = colCount):
				for cell in column:
					cell.value = columnCount
					columnCount += 1
			dataCount = 0
			for column in ws.iter_cols(min_row = 2, max_row = len(histData.data) + 1,min_col = colCount + 1, max_col = colCount + 1):
				for cell in column:
					cell.value = histData.data[dataCount]
					dataCount += 1
			colCount += 3
		startRow = maxRow + 3
		cutoffStr = str(self.limitList[0])
		del self.limitList[0]
		for limit in self.limitList:
			cutoffStr += "-" + str(limit)
		sep = "."
		proteinName = self.fileTkVar.get().split(sep, 1)[0]
		name = proteinName + "pH" + str(self.pHTkVar.get()) + 'cut' + cutoffStr 
		self.visualize()
		self.proteinImage.save(name + ".jpg")
		ws.cell(row = startRow, column = 1, value = "Protein Chain Length")
		ws.cell(row = startRow, column = 2, value = self.proteinLength)
		img = xlImage(name + ".jpg")
		ws.add_image(img, "A" + str(startRow + 3))
		wb.save(name + ".xlsx")
		
		
		#errorMessage("Successfully exported data!", 200)
	def plotDistributions(self):
		style.use(STYLE)
		font = {'fontname':'Helvetica'}
		self.proteinLength = len(self.sequenceList)
		print("protein length: ", self.proteinLength)
		if self.graphExists and not self.altGraphExists:
			self.subplot1.clear()
			self.subplot2.clear()
			#self.canvas.get_tk_widget().destroy()
			#self.toolbar.destroy()
		#print(self.sortedAminoList)
		acidID1 = int(self.graph1TkVar.get()) - 1
		acidID2 = int(self.graph2TkVar.get()) - 1
		histData1 = self.getHistogramData(acidID1)
		histData2 = self.getHistogramData(acidID2)
		print("histData2: ",histData1)
		print("histData2: ",histData2)
		if not self.graphExists	:
			self.startCanvas.destroy()
			self.plotFigure = Figure(figsize=(5.5, 3.3), dpi=100, facecolor = "#dde8f1")
			self.subplot1 = self.plotFigure.add_subplot(121)
			self.subplot2 = self.plotFigure.add_subplot(122)
			self.subplot1.set_color_cycle(COLORARRAY)
			self.subplot2.set_color_cycle(COLORARRAY)
			self.subplot1.tick_params(labelsize = 7)
			self.subplot2.tick_params(labelsize = 7)
		elif self.altGraphExists:
			self.plotFigure = Figure(figsize=(5.5, 3.3), dpi=100, facecolor = "#dde8f1")
			self.subplot1 = self.plotFigure.add_subplot(121)
			self.subplot2 = self.plotFigure.add_subplot(122)
			self.subplot1.set_color_cycle(COLORARRAY)
			self.subplot2.set_color_cycle(COLORARRAY)
			self.subplot1.tick_params(labelsize = 7)
			self.subplot2.tick_params(labelsize = 7)
		#print("histData1: ", histData1)
		#print("histData2: ", histData2)
		binwidth = 1
		if histData1 and histData2:
			maximum = max(max(histData1), max(histData2))
		elif histData1 and not histData2:
			maximum = max(histData1)
		elif not histData1 and histData2:
			maximum = max(histData2)
		else:
			maximum = 1
		hist, bins = np.histogram(histData1, bins=range(1, maximum + binwidth + 1, binwidth))
		widths = np.diff(bins)
		hist = hist / self.proteinLength
		self.subplot1.bar(bins[:-1], hist, widths, color = COLORARRAY[acidID1])
		print("hist1: ", hist)
		"""x1, y1, _ = self.subplot1.hist(histData1, color = COLORARRAY[0], normed = True, 
			bins=range(min(histData1), maximum + binwidth, binwidth))"""
		if self.numCuts == 1:
			if acidID1 == 0:
				xlabel1 = "Hydrophilic Block Size"
			else: 
				xlabel1 = "Hydrophobic Block Size"
		else:
			if acidID1 == self.numCuts:
				xlabel1 = str(self.limitList[acidID1 - 1]) + " to " + str(100) + " Hydrophobicity Block Size"
			elif acidID1 == 0:
				xlabel1 = str(-100) + " to " + str(self.limitList[acidID1]) + " Hydrophobicity Block Size"
			else:
				xlabel1 = str(self.limitList[acidID1 - 1]) + " to " + str(self.limitList[acidID1]) + " Hydrophobicity Block Size"
		self.subplot1.set_ylabel("Normalized Frequency", labelpad=5, fontsize = 8, **font)
		self.subplot1.set_xlabel(xlabel1 , labelpad = 5, fontsize = 8, **font)
		#ylim1 = self.subplot1.get_ylim()
		"""self.subplot2.hist(histData2, color = COLORARRAY[1], normed = True, 
			bins=range(min(histData2), maximum + binwidth, binwidth))"""
		hist, bins = np.histogram(histData2, bins=range(1, maximum + binwidth + 1, binwidth))
		widths = np.diff(bins)
		print("hist2: ", hist)
		print("maximum: ", maximum)
		hist = hist / self.proteinLength
		self.subplot2.bar(bins[:-1], hist, widths, color = COLORARRAY[acidID2])
		print("acidID2: ", acidID2)
		print("numCuts: ", self.numCuts)
		if self.numCuts == 1:
			if acidID2 == 1:
				xlabel2 = "Hydrophobic Block Size"
			else:
				xlabel2 = "Hydrophilic Block Size"
		else:
			if acidID2 == self.numCuts:
				xlabel2 = str(self.limitList[acidID2 - 1]) + " to " + str(100) + " Hydrophobicity Block Size"
			elif acidID2 == 0:
				xlabel2 = str(-100) + " to " + str(self.limitList[acidID2 + 1]) + " Hydrophobicity Block Size"
			else:
				xlabel2 = str(self.limitList[acidID2 - 1]) + " to " + str(self.limitList[acidID2]) + " Hydrophobicity Block Size"
		self.subplot2.set_ylabel("Normalized Frequency", labelpad=5, fontsize = 8, **font)
		self.subplot2.set_xlabel(xlabel2 , labelpad = 5, fontsize = 8, **font)
		"""ylim2 = self.subplot2.get_ylim()
		maxLim = max(max(ylim1), max(ylim2))
		print("ylim1: ", ylim1)
		print("ylim2: ", ylim2)
		self.subplot1.set_ylim([0, maxLim])
		self.subplot2.set_ylim([0, maxLim])"""
		self.plotFigure.tight_layout()
		# A tk.DrawingArea
		#imbedding matplotlib graph onto canvas
		if not self.graphExists:
			self.canvas = FigureCanvasTkAgg(self.plotFigure, master = self.canvasFrame)
		else:
			self.canvas.draw()
		self.canvas.show()
		
		#Imbedding matplotlib toolbar onto canvas
		if not self.graphExists:
			#self.toolbar = NavigationToolbar2TkAgg(self.canvas, root)
			#self.toolbar.update()
			self.canvas._tkcanvas.pack(side = Tk.BOTTOM, fill = Tk.BOTH, expand = 1)
			self.canvas.get_tk_widget().configure(bg = "black", bd = 0, selectbackground = "black")
			self.canvas.get_tk_widget().pack(side = Tk.BOTTOM, fill = Tk.BOTH, expand = 1)
		self.graphExists = True
		print("height: ", self.canvasFrame.winfo_height())
		print("width: ", self.canvasFrame.winfo_width())
	def unpackTkVarList(self, tkVarList):
		intList = []
		for tkVar in tkVarList:
			intList.append(int(tkVar.get()))
		return intList
	def sortAminoAcids(self):
		sortedList = []
		self.numCuts = int(self.numCutTkVar.get())
		self.pH = int(self.pHTkVar.get())
		self.limitList = self.unpackTkVarList(self.hydroLimitTkVarList)
		self.limitList = self.limitList[0:self.numCuts]
		self.limitList = sorted(set(self.limitList))
		self.numCuts = len(self.limitList)
		print("limitList: ", self.limitList)
		for aminoAcid in self.sequenceList:
			aminoAcidObj = getAminoAcid(aminoAcid)
			added = False
			limitID = 0
			for limit in self.limitList:
				if not added:
					if ALT:
						if aminoAcidObj.alt_hphob <= limit:
							sortedList.append(limitID)
							added = True
					else:
						if self.pH == 2:
							if aminoAcidObj.hphob2 <= limit:
								sortedList.append(limitID)
								added = True
						if self.pH == 7:
							if aminoAcidObj.hphob7 <= limit:
								sortedList.append(limitID)
								added = True
				limitID += 1
			if not added:
				sortedList.append(limitID)
		print(sortedList)

		return sortedList
		#returns an array of numbers for each consecutive monomer; to be used in histogram plotting
	def getHistogramData(self, acidID):
		#initializing array to be returned
		histogramData = []
		numConsecutive = 0
		#count through all polymers
		acidCount = 0
		for aminoAcid in self.sortedAminoList: 
			#print("acidCount: ", acidCount)
			if acidCount == len(self.sortedAminoList) - 1:
				if  aminoAcid == acidID:
					numConsecutive += 1
				count = 0
				while count < numConsecutive:
					histogramData.append(numConsecutive)
					count += 1
				numConsecutive = 0
				continue
			#if monomer is not consecutive and monomer before was, add consecutive number to data and reset values
			if aminoAcid != acidID and numConsecutive > 0:
				count = 0
				while count < numConsecutive:
					histogramData.append(numConsecutive)
					count += 1
				numConsecutive = 0
				acidCount += 1
				continue
			#increment consecutive counter by 1 if monomer is consecutive
			if  aminoAcid == acidID:
				numConsecutive += 1
				acidCount += 1
				continue
			acidCount += 1
			
		return histogramData
def adjust(toplevel, yRatio):
    toplevel.update_idletasks()
    w = toplevel.winfo_screenwidth()
    h = toplevel.winfo_screenheight()
    size = tuple(int(_) for _ in toplevel.geometry().split('+')[0].split('x'))
    x = w/2 - size[0]/2
    y = yRatio*h - size[1]/2
    toplevel.geometry("%dx%d+%d+%d" % (size + (x, y - 30)))
def errorMessage(message, width):
	#Toplevel parameters
	top = Tk.Toplevel()
	top.grab_set()
	top.wm_title("Error")
	top.geometry("%dx%d%+d%+d" % (width, 70, 250, 125))
	#Message
	msg = Tk.Message(master = top, text = message, width = 500)
	msg.pack(side = Tk.TOP, pady = 5)
	#OK button to exit
	exitButton = ttk.Button(master = top, text = "Ok", command = top.destroy, width = 7)
	exitButton.pack(side = Tk.TOP, pady = 5)
root = Tk.Tk()
root.style = ttk.Style()
root.style.theme_use("vista")
root.wm_title("Protein Analyzer %s" % VERSION)
app = Application(master = root)
app.mainloop()
